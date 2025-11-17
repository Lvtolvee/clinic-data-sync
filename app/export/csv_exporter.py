from __future__ import annotations
import csv
from datetime import date, datetime
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP

from app.custom_logging import get_logger
from app.db.extract import collect_patient_data
from app.utils.formatting import format_patient_data

log = get_logger(__name__)

CSV_ENCODING = "cp1251"
CSV_DELIMITER = ";"

CSV_HEADERS = [
    "Название лида", "Фамилия", "Имя", "Отчество", "Возраст пациента",
    "ФИО консультанта пациента", "Тип пациента 1", "Тип пациента 2",
    "ФИО доктора, проводившего первичный прием", "Дата первого визита",
    "Количество визитов в клинику",
    "Дата следующего приема и ФИО доктора, к кому пациент записан на прием",
    "Стоимость всех предварительных планов, руб.",
    "Стоимость всех согласованных планов, руб.",
    "Сумма оплаченных денег пациентом в клинику, руб.",
    "Процент выполнения плана, %",
    "Стадия", "Текущая стадия лечения", "Ответственный", "Филиал", "По рекомендации",
]

REPORT_HEADERS = [
    "ФИО", "Возраст пациента", "ФИО консультанта пациента", "Тип пациента 1",
    "Тип пациента 2", "ФИО доктора, проводившего первичный прием", "Дата первого визита",
    "Количество визитов в клинику",
    "Дата следующего приема и ФИО доктора, к кому пациент записан на прием",
    "Стоимость всех предварительных планов, руб.",
    "Стоимость всех согласованных планов, руб.",
    "Сумма оплаченных денег пациентом в клинику, руб.",
    "Процент выполнения плана, %",
    "Текущая стадия лечения", "Ответственный", "Филиал", "По рекомендации",
]

# Доп функции

def calculate_age(birth_date) -> str:
    if not birth_date:
        return "—"
    try:
        if isinstance(birth_date, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                try:
                    birth_date = datetime.strptime(birth_date, fmt).date()
                    break
                except Exception:
                    continue
        elif isinstance(birth_date, datetime):
            birth_date = birth_date.date()
        today = date.today()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return str(age)
    except Exception:
        return "—"


def normalize_date(value) -> date | None:
    if not value:
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except Exception:
                pass
    return None

#Возвращает строку даты в формате dd.MM.yyyy
def format_date_str(value) -> str:
    d = normalize_date(value)
    return d.strftime("%d.%m.%Y") if d else "—"

#Унифицированное форматирование Excel-листа
def format_excel_sheet(ws, light: bool = False):
    #Шапка
    for cell in ws[1]:
        if not light and cell.value == "Название лида":
            cell.value = "ФИО"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    #Ширина колонок
    col_widths = {
        1:40,
        2: 285,
        3: 70,
        5: 170,
        8: 140,
        9: 76,
        11: 140,
        12: 140,
        13: 140,
        14: 90,
    }
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        if col in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[col] / 6  # пересчёт из пикселей
        else:
            max_len = max(len(str(c.value)) if c.value else 0 for c in ws[get_column_letter(col)])
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    # Фиксированная высота строк
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 30

    # Выравнивание содержимого
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Формат даты dd.MM.yyyy
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "Дата первого визита":
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                if isinstance(c.value, (datetime, date)):
                    c.number_format = "DD.MM.YYYY"

    # Фильтр и закрепление 
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"

#Формирует строку данных для CSV/Excel
def convert_patient_data_to_csv_row(data: Dict[str, Any]) -> Dict[str, Any]:
    # Возраст
    age_str = calculate_age(data.get("Дата рождения"))
    age = int(age_str) if str(age_str).isdigit() else None

    # Дата первого визита
    first_visit_date = normalize_date(data.get("Дата первичного приёма"))

    # Предстоящие приёмы
    future_appointments = data.get("Предстоящие приёмы", [])
    canceled_exists = bool(future_appointments) and all(
        a.get("Статус") == "ОТМЕНЕН" for a in future_appointments
    )

    next_appointment = "—"
    for appt in future_appointments:
        if appt.get("Статус") == "ОЖИДАЕТСЯ":
            date_str = format_date_str(appt.get("Дата", appt.get("WORK_DATE_STR")))
            filial = appt.get("Филиал") or appt.get("FILIAL_NAME", "")
            doctor = appt.get("Доктор") or appt.get("DOCTOR_NAME", "")
            comment = appt.get("Комментарий") or appt.get("SCHEDAPPEALS_COMMENT", "")
            next_appointment = f"{date_str}, {filial}, {doctor}, Комментарий: {comment}"
            break

    # Стоимости и процент выполнения
    prelim_cost = sum(plan.get("Итого", 0) for plan in data.get("Комплексные планы", []))
    approved_cost = sum(plan.get("Итого", 0) for plan in data.get("Согласованные планы", []))
    paid_amount = data.get("Общая оплаченная сумма по согласованным планам", 0)
    plan_percent_value = (Decimal(paid_amount) / Decimal(prelim_cost) * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP) if prelim_cost else Decimal(0)

    # Стадия
    current_stage = data.get("Текущая стадия лечения", "—")
    if current_stage in ["Санирован", "Отказ от лечения", "Подготовка к лечению"]:
        stage = "Санирован"
    elif (canceled_exists or not future_appointments) and current_stage not in ["Не готов к реализации плана лечения",
                                                                                "Лечение в условиях медикаментозного сна",
                                                                                "Направлен в отделение профилактики на гигиену полости рта"]:
        stage = "Нет записей"
    else:
        stage = current_stage


    # Ответственный
    consultant = data.get("ФИО консультанта", "—")
    first_doctor = data.get("Доктор первичного приёма", "—")
    response_person = "—"
    if consultant != "—":
        parts = consultant.split()
        response_person = " ".join(parts[1:]) if len(parts) > 2 else consultant
    elif first_doctor != "—":
        parts = first_doctor.split()
        response_person = " ".join(parts[:2]) if len(parts) > 2 else first_doctor

    # Финальное формирование строки
    return {
        "Название лида": data.get("ФИО", "—"),
        "Фамилия": data.get("Фамилия", "—"),
        "Имя": data.get("Имя", "—"),
        "Отчество": data.get("Отчество", "—"),
        "Возраст пациента": age or "",
        "ФИО консультанта пациента": consultant,
        "Тип пациента 1": data.get("Статус пациента", "Статус не установлен"),
        "Тип пациента 2": data.get("Тип пациента", "Статус не установлен"),
        "ФИО доктора, проводившего первичный прием": first_doctor,
        "Дата первого визита": first_visit_date,
        "Количество визитов в клинику": data.get("Количество визитов в клинику", 0),
        "Дата следующего приема и ФИО доктора, к кому пациент записан на прием": next_appointment,
        "Стоимость всех предварительных планов, руб.": round(prelim_cost, 2),
        "Стоимость всех согласованных планов, руб.": round(approved_cost, 2),
        "Сумма оплаченных денег пациентом в клинику, руб.": round(paid_amount, 2),
        "Процент выполнения плана, %": "{:g}".format(plan_percent_value),
        "Стадия": stage,
        "Текущая стадия лечения": current_stage,
        "Ответственный": response_person,
        "Филиал": data.get("Филиал", "—"),
        "По рекомендации": "Да" if data.get("По рекомендации") else "Нет",
    }


# Создаёт processed_patients.csv, processed_patients.xlsx и обновляет Управленческий отчёт
def export_patients_to_csv(conn, patient_pcodes: List[str], output_file: Path) -> bool:
    try:
        output_file.parent.mkdir(parents=True, exist_ok=True)
        csv_path = output_file.with_suffix(".csv")
        excel_path = output_file.with_suffix(".xlsx")
        management_path = output_file.parent / "Управленческий отчёт.xlsx"

        # Удаляем старые файлы CSV/Excel
        for old_file in (csv_path, excel_path):
            if old_file.exists():
                try:
                    old_file.unlink()
                    log.info(f"Старый файл удалён: {old_file}")
                except Exception as e:
                    log.warning(f"Не удалось удалить старый файл {old_file}: {e}")

        csv_rows = []
        for pcode in patient_pcodes:
            try:
                log.info(f"Обрабатываем пациента {pcode}")
                data = format_patient_data(collect_patient_data(conn, pcode))
                csv_rows.append(convert_patient_data_to_csv_row(data))
            except Exception as e:
                log.error(f"Ошибка при обработке {pcode}: {e}")

        if not csv_rows:
            log.warning("Нет данных для экспорта пациентов.")
            return False

        # CSV (каждый раз заново)
        def _csv_safe(row: Dict[str, Any]) -> Dict[str, Any]:
            r = dict(row)
            v = r.get("Дата первого визита")
            if isinstance(v, (date, datetime)):
                r["Дата первого визита"] = v.strftime("%d.%m.%Y")
            return r

        with open(csv_path, "w", newline="", encoding=CSV_ENCODING) as f:
            writer = csv.DictWriter(f, fieldnames=CSV_HEADERS, delimiter=CSV_DELIMITER)
            writer.writeheader()
            for row in csv_rows:
                writer.writerow(_csv_safe(row))
        log.info(f"Создан новый CSV-файл: {csv_path}")

        # Excel (каждый раз заново)
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт"
        ws.append(CSV_HEADERS)
        for row in csv_rows:
            ws.append([row.get(h, "") for h in CSV_HEADERS])
        format_excel_sheet(ws)
        wb.save(excel_path)
        log.info(f"Создан новый Excel-файл: {excel_path}")

        # Управленческий отчёт — накопительный
        append_to_management_report(management_path, csv_rows)
        return True

    except Exception as e:
        log.error(f"Ошибка при экспорте пациентов: {e}")
        return False

#Создает CSV с персональными данными пациентов.
def export_personal_data_to_csv(conn, patient_pcodes: List[str], output_file: Path) -> bool:
    headers = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Телефон", "Email", "Адрес"]

    try:
        output_file = output_file.with_suffix(".csv")
        output_file.parent.mkdir(parents=True, exist_ok=True)
        csv_rows = []

        for pcode in patient_pcodes:
            try:
                raw = collect_patient_data(conn, pcode)
                data = format_patient_data(raw)
                csv_rows.append({
                    "Фамилия": data.get("Фамилия", "—"),
                    "Имя": data.get("Имя", "—"),
                    "Отчество": data.get("Отчество", "—"),
                    "Дата рождения": format_date_str(data.get("Дата рождения")),
                    "Телефон": data.get("Телефон", "—"),
                    "Email": data.get("Email", "—"),
                    "Адрес": data.get("Адрес", "—"),
                })
            except Exception as e:
                log.error(f"Ошибка при обработке {pcode}: {e}")
                continue

        with open(output_file, "w", newline="", encoding=CSV_ENCODING) as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=CSV_DELIMITER)
            writer.writeheader()
            writer.writerows(csv_rows)

        log.info(f"Создан CSV с персональными данными: {output_file}")
        return True

    except Exception as e:
        log.error(f"Ошибка при экспорте персональных данных: {e}")
        return False

#Добавление данных в 'Управленческий отчёт.xlsx' без дублирования, с форматами и итогами
def append_to_management_report(path: Path, rows: List[Dict[str, str]]):

    try:
        # Загрузка / создание отчёта
        if path.exists():
            wb = load_workbook(path)
            ws = wb.active
            if ws.merged_cells.ranges:
                for merged_range in list(ws.merged_cells.ranges):
                    ws.unmerge_cells(str(merged_range))
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Управленческий отчёт"
            ws.append(["№ п/п"] + REPORT_HEADERS)

        # Очистка старых итогов
        for row_idx in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row_idx, column=2).value or "").strip().lower()
            if val.startswith("комплексный пациент"):
                ws.delete_rows(row_idx, ws.max_row - row_idx + 1)
                break
        while ws.max_row > 1:
            is_empty = all((c.value is None or str(c.value).strip() == "") for c in ws[ws.max_row])
            if is_empty:
                ws.delete_rows(ws.max_row)
            else:
                break

        # Удаляем дубликаты
        existing = {str(c.value).strip() for c in ws["B"][1:] if c.value}
        added = 0
        start_idx = ws.max_row

        for r in rows:
            fio = r.get("Название лида", "").strip()
            if not fio:
                continue

            # Ищем строку с таким же ФИО
            existing_row = None
            for row in range(2, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=2).value or "").strip()
                if cell_value == fio:
                    existing_row = row
                    break

            # Если нашли — перезаписываем всю строку новыми данными
            if existing_row:
                row_idx = existing_row
            else:
                row_idx = ws.max_row + 1
                ws.append([])  # создаём новую строку, чтобы ws.cell() мог к ней обращаться

            first_visit = r.get("Дата первого визита")
            if isinstance(first_visit, (datetime, date)):
                first_visit_fmt = first_visit
            else:
                first_visit_fmt = normalize_date(first_visit)

            num_formula = f"=SUBTOTAL(3,$B$2:B{row_idx})"

            # Записываем данные по столбцам
            data_values = [
                num_formula,
                fio,
                r.get("Возраст пациента", ""),
                r.get("ФИО консультанта пациента", ""),
                r.get("Тип пациента 1", ""),
                r.get("Тип пациента 2", ""),
                r.get("ФИО доктора, проводившего первичный прием", ""),
                first_visit_fmt,
                r.get("Количество визитов в клинику", ""),
                r.get("Дата следующего приема и ФИО доктора, к кому пациент записан на прием", ""),
                r.get("Стоимость всех предварительных планов, руб.", ""),
                r.get("Стоимость всех согласованных планов, руб.", ""),
                r.get("Сумма оплаченных денег пациентом в клинику, руб.", ""),
                r.get("Процент выполнения плана, %", ""),
                r.get("Текущая стадия лечения", ""),
                r.get("Ответственный", ""),
                r.get("Филиал", ""),
                r.get("По рекомендации", ""),
            ]

            for col_idx, value in enumerate(data_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

            added += 1

        # Шапка
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Выравнивание строк
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Ширины столбцов
        col_widths = {
            1:40,
            2:285,
            3: 110,
            5: 170,
            8: 140,
            9: 76,
            11: 140,
            12: 140,
            13: 140,
            14: 90,
        }

        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            if col in col_widths:
                ws.column_dimensions[col_letter].width = col_widths[col] / 6  # пересчёт из пикселей в Excel-ширину
            else:
                # автоширина
                max_len = max(
                    len(str(c.value)) if c.value else 0
                    for c in ws[get_column_letter(col)]
                )
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        # Формат даты для "Дата первого визита"
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == "Дата первого визита":
                for r in range(2, ws.max_row + 1):
                    c = ws.cell(row=r, column=col)
                    if isinstance(c.value, (datetime, date)):
                        c.number_format = "DD.MM.YYYY"

        # Формат даты для "Дата первого визита"
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header == "Дата первого визита":
                for r in range(2, ws.max_row + 1):
                    c = ws.cell(row=r, column=col)
                    if isinstance(c.value, (datetime, date)):
                        c.number_format = "DD.MM.YYYY"


        start_row = ws.max_row + 1
        ws.append([None] * ws.max_column)
        start_row += 1

        # диапазон данных пациентов
        first_row = 2
        last_row = start_row - 2  # последняя строка с пациентами

        col_type1 = 5  # столбец E
        col_type2 = 6  # столбец F
        col_type3 = 15 # столбец O

        # Формирование диапазонов
        rng_type1 = f"{get_column_letter(col_type1)}{first_row}:{get_column_letter(col_type1)}{last_row}"
        rng_type2 = f"{get_column_letter(col_type2)}{first_row}:{get_column_letter(col_type2)}{last_row}"
        rng_type3 = f"{get_column_letter(col_type3)}{first_row}:{get_column_letter(col_type3)}{last_row}"

        first_cell_type1 = f"{get_column_letter(col_type1)}{first_row}"
        first_cell_type2 = f"{get_column_letter(col_type2)}{first_row}"
        first_cell_type3 = f"{get_column_letter(col_type3)}{first_row}"

        # Вставляем строки блока Тип пациента 2
        block_data_type1 = [
            "Комплексный пациент",
            "Не комплексный пациент",
            "Нуждается в наркозе",
            "Статус не установлен",
        ]

        orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        row_ptr = start_row

        for value in block_data_type1:
            ws.cell(row_ptr, 2, value)
            ws.cell(row_ptr, 2).fill = orange

            # Формула подсчёта
            formula = (
                f'=SUMPRODUCT(('
                f'SUBTOTAL(103,OFFSET({first_cell_type1},ROW({rng_type1})-ROW({first_cell_type1}),0))'
                f')*({rng_type1}="{value}"))'
            )
            ws.cell(row_ptr, 3, formula)

            row_ptr += 1

        # Пустая строка между группами
        row_ptr += 1

        # Вставляем строки блока Тип пациента 1
        block_data_type2 = [
            "Санирован",
            "Отказ от лечения",
            "Готовность к реализации (Готов к реализации + готов по специализации)",
            "Думает",
            "Статус не установлен",
        ]

        for value in block_data_type2:
            ws.cell(row_ptr, 2, value)
            ws.cell(row_ptr, 2).fill = yellow

            if value.startswith("Санирован"):
                formula = (
                    f'=SUMPRODUCT(('
                    f'SUBTOTAL(103,OFFSET({first_cell_type3},ROW({rng_type3})-ROW({first_cell_type3}),0))'
                    f')*({rng_type3}="{value}"))'
                )
            elif value.startswith("Готовность"):
                # специальная формула ИЛИ
                conditions = [
                    "Готов к реализации плана лечения",
                    "Готов по специализации",
                ]
                conditions_or = "+".join([f'({rng_type2}="{c}")' for c in conditions])

                formula = (
                    f'=SUMPRODUCT(('
                    f'SUBTOTAL(103,OFFSET({first_cell_type2},ROW({rng_type2})-ROW({first_cell_type2}),0))>0)'
                    f'*(({conditions_or})))'
                )
            else:
                # обычная формула
                formula = (
                    f'=SUMPRODUCT(('
                    f'SUBTOTAL(103,OFFSET({first_cell_type2},ROW({rng_type2})-ROW({first_cell_type2}),0))'
                    f')*({rng_type2}="{value}"))'
                )

            ws.cell(row_ptr, 3, formula)
            row_ptr += 1

        # ПУСТАЯ строка перед итогами
        ws.append([])
        row_ptr += 1

        #ИТОГОВЫЙ БЛОК

        headers = [c.value for c in ws[1]]

        col_title = headers.index("ФИО") + 1
        col_next = headers.index("Дата следующего приема и ФИО доктора, к кому пациент записан на прием") + 1
        col_paid = headers.index("Сумма оплаченных денег пациентом в клинику, руб.") + 1

        rng_title = f"{get_column_letter(col_title)}{first_row}:{get_column_letter(col_title)}{last_row}"
        rng_next = f"{get_column_letter(col_next)}{first_row}:{get_column_letter(col_next)}{last_row}"
        first_cell = rng_next.split(":")[0]
        rng_paid = f"{get_column_letter(col_paid)}{first_row}:{get_column_letter(col_paid)}{last_row}"

        start = row_ptr

        ws.cell(start, 2).value = "Итоговые показатели по текущему фильтру"
        ws.cell(start, 2).font = Font(bold=True)

        def write_metric(row, name, formula):
            ws.cell(row, 2).value = name
            ws.cell(row, 3).value = f"={formula}"
            ws.cell(row, 2).font = Font(bold=True)

        write_metric(start + 1, "Количество пациентов", f"SUBTOTAL(103,{rng_title})")
        write_metric(
            start + 2,
            "Записались повторно",
            f'SUMPRODUCT((SUBTOTAL(103,OFFSET({first_cell},ROW({rng_next})-MIN(ROW({rng_next})),0)))*(({rng_next}<>"" )*({rng_next}<>"—")))'
        )
        write_metric(
            start + 3,
            "Конверсия, %",
            f"IFERROR(({get_column_letter(3)}{start + 2}/{get_column_letter(3)}{start + 1})*100,0)"
        )
        write_metric(start + 4, "Общая сумма оплат, руб.", f"SUBTOTAL(9,{rng_paid})")
        write_metric(
            start + 5,
            "Средняя стоимость лечения, руб.",
            f"IFERROR(SUBTOTAL(9,{rng_paid})/SUBTOTAL(103,{rng_title}),0)"
        )

        # Ограничиваем автофильтр только областью пациентов
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"
        ws.freeze_panes = "A2"


        wb.save(path)
        log.info(f"Добавлено {added} строк в {path}")



    except Exception as e:
        log.error(f"Ошибка при обновлении Управленческого отчёта: {e}")

